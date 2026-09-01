import requests
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

list_url = 'https://forecast.geeksoft.tech/api/v1/forecast/list'
load_url = 'https://forecast.geeksoft.tech/api/v1/forecast/load'
run_url = 'https://forecast.geeksoft.tech/api/v1/forecast/run'

res = requests.get(list_url)
scenarios = res.json()

print(f"=== INICIANDO AUDITORÍA TOTAL PERICIAL (CON SUBTOTAALES, EMPRESA Y ACUMULADO) ===")

output_dir = 'C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/QC_Auditoria_Escenarios'
os.makedirs(output_dir, exist_ok=True)

# Paleta canónica ARGB
PALETTE = {
    'SPCC': 'FF0369A1',         # Sky 700
    'NEXA': 'FF0F4C81',         # Petral Blue
    'SPOT': 'FFF97316',         # Orange
    'MATARANI': 'FF06B6D4',     # Cyan 500
    'MARCONA': 'FFA855F7',      # Purple 500
    'MEJILLONES': 'FFD946EF',   # Fuchsia 500
    'TABLONES': 'FFDC2626',     # Red 600
    'MOQUEGUA': 'FF16A34A',     # Green 600
    'CONCON': 'FF475569',       # Slate 600
    'HUEMUL': 'FF4F46E5',       # Indigo 600
    'HEADER': 'FF1E293B',       # Slate 800
    'HEADER_TOT': 'FF0C4A6E',   # Sky 900
    'SUBTOTAL_DIM': 'FF1E293B', # Slate 800
    'SUBTOTAL_DATA': 'FFFFFBEB',# Amber 50
    'FLOTA_DIM': 'FF1E293B',    # Slate 800
    'FLOTA_DATA': 'FFF1F5F9',   # Slate 100
    'ACUM_DIM': 'FF0D9488',     # Teal 600
    'ACUM_DATA': 'FFEEF2FF',    # Indigo 50
}

audit_summary = []

for idx, sc in enumerate(scenarios, 1):
    sc_id = sc['id']
    sc_name = sc['name']
    print(f"\n[{idx}/{len(scenarios)}] Procesando: '{sc_name}'...")

    load_res = requests.get(f"{load_url}/{sc_id}")
    sc_data = load_res.json()

    s_date = sc_data.get('start_date', '2027-01-01')
    e_date = sc_data.get('end_date', '2027-12-31')
    p_lines = sc_data.get('projection_lines', [])

    if not p_lines:
        continue

    sim_res = requests.post(run_url, json={
        'start_date': s_date,
        'end_date': e_date,
        'projection_lines': p_lines
    })

    if sim_res.status_code != 200:
        continue

    sim_data = sim_res.json()
    agg_data = sim_data.get('aggregated_data', {})
    clients = list(agg_data.keys())

    # Meses YYYY-MM
    s_year, s_month = map(int, s_date.split('-')[:2])
    e_year, e_month = map(int, e_date.split('-')[:2])
    month_keys = []
    month_labels = []
    cy, cm = s_year, s_month
    month_names = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Set', 'Oct', 'Nov', 'Dic']
    while cy < e_year or (cy == e_year and cm <= e_month):
        month_keys.append(f"{cy}-{cm:02d}")
        month_labels.append(f"{month_names[cm-1]} {cy}")
        cm += 1
        if cm > 12:
            cm = 1
            cy += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matriz Financiera'
    ws.views.sheetView[0].showGridLines = True

    # 1. Cabecera
    header_cols = ['CLIENTE', 'RUTA', 'BUQUE', 'MÉTRICA'] + month_labels + ['TOTAL ACUM']
    ws.append(header_cols)
    ws.row_dimensions[1].height = 25

    slate_fill = PatternFill(start_color=PALETTE['HEADER'], end_color=PALETTE['HEADER'], fill_type='solid')
    sky_fill = PatternFill(start_color=PALETTE['HEADER_TOT'], end_color=PALETTE['HEADER_TOT'], fill_type='solid')
    white_bold = Font(name='Segoe UI', size=9, bold=True, color='FFFFFFFF')
    amber_bold = Font(name='Segoe UI', size=9, bold=True, color='FFFBBF24')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='FFE2E8F0'),
        right=Side(style='thin', color='FFE2E8F0'),
        top=Side(style='thin', color='FFE2E8F0'),
        bottom=Side(style='thin', color='FFE2E8F0')
    )

    for c_idx in range(1, len(header_cols) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = sky_fill if c_idx == len(header_cols) else slate_fill
        cell.font = white_bold
        cell.alignment = center_align
        cell.border = thin_border

    current_row = 2

    # Acumuladores Globales de Flota
    fleet_trips = [0] * len(month_keys)
    fleet_days = [0] * len(month_keys)
    fleet_tons = [0] * len(month_keys)
    fleet_net_rev = [0] * len(month_keys)
    fleet_hire = [0] * len(month_keys)
    fleet_bunker = [0] * len(month_keys)
    fleet_port = [0] * len(month_keys)
    fleet_pl = [0] * len(month_keys)

    for client in clients:
        routes_data = agg_data[client]
        client_start_row = current_row

        # Acumuladores de Subtotal por Cliente
        cl_trips = [0] * len(month_keys)
        cl_days = [0] * len(month_keys)
        cl_tons = [0] * len(month_keys)
        cl_net_rev = [0] * len(month_keys)
        cl_hire = [0] * len(month_keys)
        cl_bunker = [0] * len(month_keys)
        cl_port = [0] * len(month_keys)
        cl_pl = [0] * len(month_keys)

        for route in routes_data.keys():
            vessels_data = routes_data[route]
            route_start_row = current_row

            # Color de ruta exacto
            route_color = PALETTE['MARCONA']
            if 'MATARANI' in route.upper(): route_color = PALETTE['MATARANI']
            elif 'MEJILLONES' in route.upper(): route_color = PALETTE['MEJILLONES']
            elif 'SPOT' in route.upper(): route_color = PALETTE['SPOT']

            for vessel in vessels_data.keys():
                node = vessels_data[vessel]
                vessel_start_row = current_row

                # Color de buque exacto
                vessel_color = PALETTE['MOQUEGUA']
                if 'TABLONES' in vessel.upper(): vessel_color = PALETTE['TABLONES']
                elif 'CONCON' in vessel.upper(): vessel_color = PALETTE['CONCON']
                elif 'HUEMUL' in vessel.upper(): vessel_color = PALETTE['HUEMUL']

                trips_series = []
                days_series = []
                tons_series = []
                net_rev_series = []
                hire_series = []
                bunker_series = []
                port_series = []
                pl_series = []
                tce_series = []

                for m_idx, m_key in enumerate(month_keys):
                    md = node.get(m_key, {})
                    t = md.get('freq', 0)
                    d = md.get('total_duration', 0) if t > 0 else 0
                    c = md.get('carga_unit', 13500) * t if t > 0 else 0
                    nr = md.get('net_income', 0) if t > 0 else 0
                    h = md.get('charter_hire_cost', 0) if t > 0 else 0
                    b = md.get('total_bunker_costs', 0) if t > 0 else 0
                    p = md.get('total_port_costs', 0) if t > 0 else 0
                    pl = md.get('voyage_result', 0) if t > 0 else 0
                    tce = md.get('tce_real', 0) if t > 0 else 0

                    trips_series.append(t)
                    days_series.append(d)
                    tons_series.append(c)
                    net_rev_series.append(nr)
                    hire_series.append(h)
                    bunker_series.append(b)
                    port_series.append(p)
                    pl_series.append(pl)
                    tce_series.append(tce)

                    # Sumar a cliente y flota
                    cl_trips[m_idx] += t
                    cl_days[m_idx] += d
                    cl_tons[m_idx] += c
                    cl_net_rev[m_idx] += nr
                    cl_hire[m_idx] += h
                    cl_bunker[m_idx] += b
                    cl_port[m_idx] += p
                    cl_pl[m_idx] += pl

                    fleet_trips[m_idx] += t
                    fleet_days[m_idx] += d
                    fleet_tons[m_idx] += c
                    fleet_net_rev[m_idx] += nr
                    fleet_hire[m_idx] += h
                    fleet_bunker[m_idx] += b
                    fleet_port[m_idx] += p
                    fleet_pl[m_idx] += pl

                sum_fn = lambda l: sum(l)
                avg_fn = lambda l: (sum(l) / len([x for x in l if x > 0])) if any(x > 0 for x in l) else 0

                node_metrics = [
                    ('Viajes (freq)', trips_series, sum_fn(trips_series), '0.0', False),
                    ('Días-Buque', days_series, sum_fn(days_series), '0.0', False),
                    ('Toneladas', tons_series, sum_fn(tons_series), '#,##0', False),
                    ('Net Revenue', net_rev_series, sum_fn(net_rev_series), '$#,##0', True),
                    ('(-) Hire (TCE x días)', hire_series, sum_fn(hire_series), '$#,##0', False),
                    ('(-) Bunker Costs', bunker_series, sum_fn(bunker_series), '$#,##0', False),
                    ('(-) Port Costs', port_series, sum_fn(port_series), '$#,##0', False),
                    ('(=) VOYAGE RESULT / P&L', pl_series, sum_fn(pl_series), '$#,##0', True),
                    ('Métricas TCE ($/d)', tce_series, avg_fn(tce_series), '$#,##0.00', False),
                ]

                for m_name, vals, tot_val, num_fmt, is_bold in node_metrics:
                    clean_vals = [v if (v is not None and v != 0) else '' for v in vals]
                    clean_tot = tot_val if (tot_val is not None and tot_val != 0) else ''
                    row_vals = [
                        client if current_row == client_start_row else '',
                        route if current_row == route_start_row else '',
                        vessel if current_row == vessel_start_row else '',
                        m_name
                    ] + clean_vals + [clean_tot]
                    ws.append(row_vals)
                    ws.row_dimensions[current_row].height = 19

                    for c_idx in range(5, len(header_cols) + 1):
                        cell = ws.cell(row=current_row, column=c_idx)
                        if cell.value != '':
                            cell.number_format = num_fmt
                        cell.font = Font(name='Segoe UI', size=8.5, bold=is_bold or c_idx == len(header_cols))
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.border = thin_border

                    m_cell = ws.cell(row=current_row, column=4)
                    m_cell.font = Font(name='Segoe UI', size=8.5, bold=is_bold)
                    m_cell.alignment = Alignment(horizontal='left', vertical='center')
                    m_cell.border = thin_border

                    current_row += 1

                # Merge de Buque
                v_end_row = current_row - 1
                if v_end_row > vessel_start_row:
                    ws.merge_cells(start_row=vessel_start_row, start_column=3, end_row=v_end_row, end_column=3)
                v_cell = ws.cell(row=vessel_start_row, column=3)
                v_cell.fill = PatternFill(start_color=vessel_color, end_color=vessel_color, fill_type='solid')
                v_cell.font = white_bold
                v_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

            # Merge de Ruta
            r_end_row = current_row - 1
            if r_end_row > route_start_row:
                ws.merge_cells(start_row=route_start_row, start_column=2, end_row=r_end_row, end_column=2)
            r_cell = ws.cell(row=route_start_row, column=2)
            r_cell.fill = PatternFill(start_color=route_color, end_color=route_color, fill_type='solid')
            r_cell.font = white_bold
            r_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

        # --- FILAS DE SUBTOTAL CLIENTE ---
        subtotal_start_row = current_row
        cl_metrics = [
            ('Viajes', cl_trips, sum(cl_trips), '0.0'),
            ('Días-Buque', cl_days, sum(cl_days), '0.0'),
            ('Toneladas', cl_tons, sum(cl_tons), '#,##0'),
            ('Net Revenue', cl_net_rev, sum(cl_net_rev), '$#,##0'),
            ('(-) Hire (TCE x días)', cl_hire, sum(cl_hire), '$#,##0'),
            ('(-) Bunker Costs', cl_bunker, sum(cl_bunker), '$#,##0'),
            ('(-) Port Costs', cl_port, sum(cl_port), '$#,##0'),
            ('(=) VOYAGE RESULT / P&L', cl_pl, sum(cl_pl), '$#,##0'),
        ]

        for s_idx, (m_name, vals, tot_val, num_fmt) in enumerate(cl_metrics):
            clean_vals = [v if (v is not None and v != 0) else '' for v in vals]
            clean_tot = tot_val if (tot_val is not None and tot_val != 0) else ''
            row_vals = [
                client if current_row == client_start_row else '',
                'Σ SUBTOTAL' if s_idx == 0 else '',
                'TOTAL CLIENT' if s_idx == 0 else '',
                m_name
            ] + clean_vals + [clean_tot]
            ws.append(row_vals)
            ws.row_dimensions[current_row].height = 20

            # Celdas de datos con fondo Amber 50
            for c_idx in range(5, len(header_cols) + 1):
                cell = ws.cell(row=current_row, column=c_idx)
                if cell.value != '':
                    cell.number_format = num_fmt
                cell.font = Font(name='Segoe UI', size=8.5, bold=True, color='FF78350F')
                cell.fill = PatternFill(start_color=PALETTE['SUBTOTAL_DATA'], end_color=PALETTE['SUBTOTAL_DATA'], fill_type='solid')
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = thin_border

            m_cell = ws.cell(row=current_row, column=4)
            m_cell.font = Font(name='Segoe UI', size=8.5, bold=True, color='FF78350F')
            m_cell.fill = PatternFill(start_color=PALETTE['SUBTOTAL_DATA'], end_color=PALETTE['SUBTOTAL_DATA'], fill_type='solid')
            m_cell.alignment = Alignment(horizontal='left', vertical='center')
            m_cell.border = thin_border

            current_row += 1

        subtotal_end_row = current_row - 1
        # Merge de Subtotal Ruta + Buque (Cols 2 y 3)
        ws.merge_cells(start_row=subtotal_start_row, start_column=2, end_row=subtotal_end_row, end_column=3)
        sub_dim_cell = ws.cell(row=subtotal_start_row, column=2)
        sub_dim_cell.value = 'Σ SUBTOTAL\nTOTAL CLIENT'
        sub_dim_cell.fill = PatternFill(start_color=PALETTE['SUBTOTAL_DIM'], end_color=PALETTE['SUBTOTAL_DIM'], fill_type='solid')
        sub_dim_cell.font = amber_bold
        sub_dim_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Merge de Cliente (incluyendo sus subtotales)
        c_end_row = current_row - 1
        if c_end_row > client_start_row:
            ws.merge_cells(start_row=client_start_row, start_column=1, end_row=c_end_row, end_column=1)
        c_cell = ws.cell(row=client_start_row, column=1)
        c_color = PALETTE['SPCC'] if client == 'SPCC' else PALETTE['NEXA']
        c_cell.fill = PatternFill(start_color=c_color, end_color=c_color, fill_type='solid')
        c_cell.font = white_bold
        c_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

    # --- 3. BLOQUE TOTAL FLOTA (EMPRESA) ---
    fleet_start_row = current_row
    fleet_metrics = [
        ('Viajes', fleet_trips, sum(fleet_trips), '0.0'),
        ('Días-Buque', fleet_days, sum(fleet_days), '0.0'),
        ('Toneladas', fleet_tons, sum(fleet_tons), '#,##0'),
        ('Net Revenue', fleet_net_rev, sum(fleet_net_rev), '$#,##0'),
        ('(-) Hire (TCE x días)', fleet_hire, sum(fleet_hire), '$#,##0'),
        ('(-) Bunker Costs', fleet_bunker, sum(fleet_bunker), '$#,##0'),
        ('(-) Port Costs', fleet_port, sum(fleet_port), '$#,##0'),
        ('(=) VOYAGE RESULT / P&L', fleet_pl, sum(fleet_pl), '$#,##0'),
    ]

    for f_idx, (m_name, vals, tot_val, num_fmt) in enumerate(fleet_metrics):
        clean_vals = [v if (v is not None and v != 0) else '' for v in vals]
        clean_tot = tot_val if (tot_val is not None and tot_val != 0) else ''
        row_vals = [
            'TOTAL FLOTA' if f_idx == 0 else '',
            '',
            '',
            m_name
        ] + clean_vals + [clean_tot]
        ws.append(row_vals)
        ws.row_dimensions[current_row].height = 20

        for c_idx in range(5, len(header_cols) + 1):
            cell = ws.cell(row=current_row, column=c_idx)
            if cell.value != '':
                cell.number_format = num_fmt
            cell.font = Font(name='Segoe UI', size=8.5, bold=True, color='FF1E293B')
            cell.fill = PatternFill(start_color=PALETTE['FLOTA_DATA'], end_color=PALETTE['FLOTA_DATA'], fill_type='solid')
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = thin_border

        m_cell = ws.cell(row=current_row, column=4)
        m_cell.font = Font(name='Segoe UI', size=8.5, bold=True, color='FF1E293B')
        m_cell.fill = PatternFill(start_color=PALETTE['FLOTA_DATA'], end_color=PALETTE['FLOTA_DATA'], fill_type='solid')
        m_cell.alignment = Alignment(horizontal='left', vertical='center')
        m_cell.border = thin_border

        current_row += 1

    fleet_end_row = current_row - 1
    # Merge Cols 1, 2, 3 para TOTAL FLOTA
    ws.merge_cells(start_row=fleet_start_row, start_column=1, end_row=fleet_end_row, end_column=3)
    fleet_cell = ws.cell(row=fleet_start_row, column=1)
    fleet_cell.value = 'TOTAL FLOTA\n(EMPRESA)'
    fleet_cell.fill = PatternFill(start_color=PALETTE['FLOTA_DIM'], end_color=PALETTE['FLOTA_DIM'], fill_type='solid')
    fleet_cell.font = white_bold
    fleet_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- 4. BLOQUE TOTAL ACUMULADO GLOBAL ---
    acum_start_row = current_row
    accum_fn = lambda arr: [sum(arr[:i+1]) for i in range(len(arr))]
    
    acum_trips = accum_fn(fleet_trips)
    acum_days = accum_fn(fleet_days)
    acum_tons = accum_fn(fleet_tons)
    acum_net_rev = accum_fn(fleet_net_rev)
    acum_hire = accum_fn(fleet_hire)
    acum_bunker = accum_fn(fleet_bunker)
    acum_port = accum_fn(fleet_port)
    acum_pl = accum_fn(fleet_pl)

    acum_metrics = [
        ('Viajes', acum_trips, acum_trips[-1] if acum_trips else 0, '0.0'),
        ('Días-Buque', acum_days, acum_days[-1] if acum_days else 0, '0.0'),
        ('Toneladas', acum_tons, acum_tons[-1] if acum_tons else 0, '#,##0'),
        ('Net Revenue', acum_net_rev, acum_net_rev[-1] if acum_net_rev else 0, '$#,##0'),
        ('(-) Hire (TCE x días)', acum_hire, acum_hire[-1] if acum_hire else 0, '$#,##0'),
        ('(-) Bunker Costs', acum_bunker, acum_bunker[-1] if acum_bunker else 0, '$#,##0'),
        ('(-) Port Costs', acum_port, acum_port[-1] if acum_port else 0, '$#,##0'),
        ('(=) VOYAGE RESULT / P&L', acum_pl, acum_pl[-1] if acum_pl else 0, '$#,##0'),
    ]

    for a_idx, (m_name, vals, tot_val, num_fmt) in enumerate(acum_metrics):
        clean_vals = [v if (v is not None and v != 0) else '' for v in vals]
        clean_tot = tot_val if (tot_val is not None and tot_val != 0) else ''
        row_vals = [
            'TOTAL ACUMULADO' if a_idx == 0 else '',
            '',
            '',
            m_name
        ] + clean_vals + [clean_tot]
        ws.append(row_vals)
        ws.row_dimensions[current_row].height = 20

        for c_idx in range(5, len(header_cols) + 1):
            cell = ws.cell(row=current_row, column=c_idx)
            if cell.value != '':
                cell.number_format = num_fmt
            cell.font = Font(name='Segoe UI', size=8.5, bold=True, color='FF1E1B4B')
            cell.fill = PatternFill(start_color=PALETTE['ACUM_DATA'], end_color=PALETTE['ACUM_DATA'], fill_type='solid')
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = thin_border

        m_cell = ws.cell(row=current_row, column=4)
        m_cell.font = Font(name='Segoe UI', size=8.5, bold=True, color='FF1E1B4B')
        m_cell.fill = PatternFill(start_color=PALETTE['ACUM_DATA'], end_color=PALETTE['ACUM_DATA'], fill_type='solid')
        m_cell.alignment = Alignment(horizontal='left', vertical='center')
        m_cell.border = thin_border

        current_row += 1

    acum_end_row = current_row - 1
    # Merge Cols 1, 2, 3 para TOTAL ACUMULADO
    ws.merge_cells(start_row=acum_start_row, start_column=1, end_row=acum_end_row, end_column=3)
    acum_cell = ws.cell(row=acum_start_row, column=1)
    acum_cell.value = 'TOTAL ACUMULADO\n(CIERRE)'
    acum_cell.fill = PatternFill(start_color=PALETTE['ACUM_DIM'], end_color=PALETTE['ACUM_DIM'], fill_type='solid')
    acum_cell.font = white_bold
    acum_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 30
    for col_letter in [openpyxl.utils.get_column_letter(i) for i in range(5, len(header_cols) + 1)]:
        ws.column_dimensions[col_letter].width = 16

    safe_filename = "".join(c for c in sc_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    file_path = f"{output_dir}/Matriz_{safe_filename}.xlsx"
    wb.save(file_path)

    audit_summary.append({
        'name': sc_name,
        'clients': clients,
        'rows': current_row - 1,
        'cols': len(header_cols),
        'trips': sum(fleet_trips),
        'net_revenue': sum(fleet_net_rev),
        'voyage_margin': sum(fleet_pl),
        'file': file_path
    })

print("\n=== RESUMEN DE AUDITORÍA TOTAL COMPLETADA ===")
for res in audit_summary:
    print(f" -> Escenario: {res['name']}")
    print(f"    Total Filas (con Subtotales + Empresa + Acumulado): {res['rows']}")
    print(f"    Total Viajes: {res['trips']} | Net Revenue: ${res['net_revenue']:,.0f} | P/L: ${res['voyage_margin']:,.0f}")
    print(f"    Archivo: {res['file']}")
