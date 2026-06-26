import openpyxl
import psycopg2
import sys

# 1. Definir los archivos Excel a procesar
files = {
    "TABLONES": "C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Tablones.xlsx",
    "MOQUEGUA": "C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Moquegua.xlsx",
    "CONCON_TRADER": "C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Concon_Trader.xlsx"
}

# 2. Las métricas base ya conciliadas y extraídas previamente
excel_scenarios = [
    {"vessel": "TABLONES", "route": "ILO-MATARANI", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 0.5384, "bunker": 20601.88, "voy_res": 195033.12, "tot_dur": 4.0801, "tce_real": 47801.35, "pl_vs_req": 133831.98},
    {"vessel": "TABLONES", "route": "ILO-MARCONA", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 2.2083, "bunker": 42275.73, "voy_res": 198794.27, "tot_dur": 5.7499, "tce_real": 34573.37, "pl_vs_req": 112545.40},
    {"vessel": "TABLONES", "route": "ILO-MEJILLONES", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 2.6140, "bunker": 50042.28, "voy_res": 176702.72, "tot_dur": 6.1557, "tce_real": 28705.63, "pl_vs_req": 84367.50},
    {"vessel": "MOQUEGUA", "route": "ILO-MATARANI", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 0.5384, "bunker": 18560.53, "voy_res": 199074.47, "tot_dur": 4.0801, "tce_real": 48791.86, "pl_vs_req": 146033.49},
    {"vessel": "MOQUEGUA", "route": "ILO-MARCONA", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 2.2083, "bunker": 39487.00, "voy_res": 206583.00, "tot_dur": 5.7499, "tce_real": 35927.95, "pl_vs_req": 131833.98},
    {"vessel": "MOQUEGUA", "route": "ILO-MEJILLONES", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 2.6140, "bunker": 47071.94, "voy_res": 183673.06, "tot_dur": 6.1557, "tce_real": 29837.97, "pl_vs_req": 103649.20},
    {"vessel": "CONCON_TRADER", "route": "ILO-MATARANI", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 0.5384, "bunker": 20360.91, "voy_res": 193774.09, "tot_dur": 4.0801, "tce_real": 47492.77, "pl_vs_req": 112172.58},
    {"vessel": "CONCON_TRADER", "route": "ILO-MARCONA", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 2.2083, "bunker": 41287.38, "voy_res": 182282.62, "tot_dur": 5.7499, "tce_real": 31701.74, "pl_vs_req": 67284.13},
    {"vessel": "CONCON_TRADER", "route": "ILO-MEJILLONES", "act_load": 500, "act_disch": 300, "port_days": 3.5417, "sea_days": 2.6140, "bunker": 48872.32, "voy_res": 137372.68, "tot_dur": 6.1557, "tce_real": 22316.40, "pl_vs_req": 14259.04}
]

# Mapa heurístico para encontrar la hoja correcta en cada Excel según la ruta
def match_sheet(route, sheet_names):
    dest = route.split('-')[1].upper()
    for sn in sheet_names:
        if dest in sn.upper() and 'RESUMEN' not in sn.upper():
            return sn
    return None

def extract_additional_expenses(filepath, route):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_name = match_sheet(route, wb.sheetnames)
        if not sheet_name:
            return 0
        
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if cell and isinstance(cell, str) and 'additional expenses' in cell.strip().lower():
                    # Buscar el numero adyacente (a la derecha)
                    if i + 1 < len(row):
                        val = row[i+1]
                        if isinstance(val, (int, float)):
                            return float(val)
        return 0
    except Exception as e:
        print(f"  [!] Error procesando {filepath}: {e}")
        return 0

# 3. Extraer valores faltantes dinámicamente
print("Iniciando Extraccion desde Exceles...")
for s in excel_scenarios:
    filepath = files.get(s["vessel"])
    if not filepath: continue
    
    additional_expenses = extract_additional_expenses(filepath, s["route"])
    s["additional_expenses"] = additional_expenses
    print(f"  {s['vessel']} -> {s['route']} | Additional Expenses: ${additional_expenses}")

# 4. Guardar en Supabase
conn_str = "postgresql://postgres.hjjxooxcpvlvbaxgifbn:VivaLaVida2026$@aws-1-us-east-2.pooler.supabase.com:6543/postgres"

print("\nSubiendo metricas a Supabase (audit_benchmarks)...")
try:
    conn = psycopg2.connect(conn_str)
    conn.autocommit = True
    cur = conn.cursor()

    # Actualizamos el registro usando UPSERT simulado por UPDATE o borramos y reinsertamos
    cur.execute("TRUNCATE TABLE public.audit_benchmarks;")
    
    insert_query = """
        INSERT INTO public.audit_benchmarks 
        (scenario_key, act_load, act_disch, port_days, sea_days, bunker_costs, voyage_result, total_duration, tce_real, pl_vs_req, additional_expenses)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    for s in excel_scenarios:
        key = f"{s['vessel']}-{s['route']}"
        cur.execute(insert_query, (
            key, s['act_load'], s['act_disch'], s['port_days'], s['sea_days'], 
            s['bunker'], s['voy_res'], s['tot_dur'], s['tce_real'], s['pl_vs_req'], s['additional_expenses']
        ))
        
    print("¡9 Escenarios cargados a DB exitosamente!")

except Exception as e:
    print(f"Error DB: {e}")
finally:
    if 'cur' in locals(): cur.close()
    if 'conn' in locals(): conn.close()
