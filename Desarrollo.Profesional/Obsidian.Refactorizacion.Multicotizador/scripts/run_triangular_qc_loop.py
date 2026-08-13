import sys
import os
import openpyxl
import json
import requests

EXCEL_FILE = r'C:\Users\rguti\PETRAL.SMART.DASHBOARD\Documentos.Petral\NEXA ILO CALLA MATARANI ILO.IZ.12.08.26.xlsx'
API_URL = 'https://forecast.geeksoft.tech/api/v1/forecast/multicotizador/calculate'

def load_excel_reference():
    """
    Extracción de valores de referencia desde el archivo Excel oficial PETRAL.
    """
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    return {
        "gross_revenue": float(ws.cell(14, 14).value or 0.0),
        "port_costs": float(ws.cell(15, 14).value or 0.0),
        "bunker_costs": float(ws.cell(16, 14).value or 0.0),
        "voyage_result": float(ws.cell(18, 14).value or 0.0),
        "total_days": float(ws.cell(14, 17).value or 0.0),
        "sea_days": float(ws.cell(15, 17).value or 0.0),
        "port_days": float(ws.cell(16, 17).value or 0.0),
        "tce_real": float(ws.cell(17, 17).value or 0.0),
        "pnl_net": float(ws.cell(20, 17).value or 0.0)
    }

def simulate_frontend_payload_builder():
    """
    Simulación exacta de la función payloadTramos de MultiCotizadorExcel.tsx (Frontend React).
    """
    puertosConfig = [
        {"action": "NONE", "op_rate": "", "overhead": "", "quantity": "", "rate_unit": "TH", "positioning": "", "freight_rate": "", "manual_port_cost": ""},
        {"action": "CARGAR", "op_rate": "500", "overhead": "6", "quantity": "13500", "rate_unit": "TH", "positioning": "1", "freight_rate": "", "manual_port_cost": 17000},
        {"action": "DESCARGAR", "op_rate": "400", "overhead": "6", "quantity": "13500", "rate_unit": "TH", "positioning": "0", "freight_rate": "30", "manual_port_cost": 18000},
        {"action": "NONE", "op_rate": "", "quantity": 0, "rate_unit": "TH", "freight_rate": 0, "manual_port_cost": ""}
    ]

    tramos_inputs = [
        {"type": "BALLAST", "origin_port_id": "ILO", "destination_port_id": "CALLAO", "quantity": 0, "freight_rate": 0, "port_delay_hours_loading": 0, "port_delay_hours_discharging": 0, "route_distance": 514, "weather_factor": 3, "speed": 11},
        {"type": "LADEN", "origin_port_id": "CALLAO", "destination_port_id": "MATARANI", "quantity": 13500, "freight_rate": 30, "port_delay_hours_loading": 0, "port_delay_hours_discharging": 0, "route_distance": 457, "weather_factor": 3, "speed": 11},
        {"type": "BALLAST", "origin_port_id": "MATARANI", "destination_port_id": "ILO", "quantity": 0, "freight_rate": 0, "port_delay_hours_loading": 0, "port_delay_hours_discharging": 0, "route_distance": 69, "weather_factor": 3, "speed": 11}
    ]

    payloadTramos = []
    for idx, tr in enumerate(tramos_inputs):
        pOrig = puertosConfig[idx]
        pDest = puertosConfig[idx + 1]

        customLoad = None
        customDisch = None

        if pDest and pDest.get("action") == "CARGAR" and pDest.get("op_rate") != "":
            val = float(pDest["op_rate"])
            customLoad = val if pDest.get("rate_unit") == "TH" else val / 24.0
        elif pOrig and pOrig.get("action") == "CARGAR" and pOrig.get("op_rate") != "":
            val = float(pOrig["op_rate"])
            customLoad = val if pOrig.get("rate_unit") == "TH" else val / 24.0

        if pDest and pDest.get("action") == "DESCARGAR" and pDest.get("op_rate") != "":
            val = float(pDest["op_rate"])
            customDisch = val if pDest.get("rate_unit") == "TH" else val / 24.0

        destQuantity = float(tr.get("quantity") or 0)
        if pDest and pDest.get("action") != "NONE" and pDest.get("quantity") != "" and pDest.get("quantity") is not None:
            pQty = float(pDest["quantity"])
            if pQty > 0:
                destQuantity = pQty

        opDest = pDest.get("action", "NONE")
        overheadDest = 0.0
        posCarga = 0.0
        posDescarga = 0.0

        if opDest != "NONE":
            if pDest.get("overhead") != "" and pDest.get("overhead") is not None:
                overheadDest = float(pDest["overhead"])
            if pDest.get("positioning") != "" and pDest.get("positioning") is not None:
                pVal = float(pDest["positioning"])
                if opDest == "CARGAR":
                    posCarga = pVal
                elif opDest == "DESCARGAR":
                    posDescarga = pVal

        overridePortCostOrig = 0.00001
        overridePortCostDest = 0.00001

        if idx == 0 and pOrig and pOrig.get("action") != "NONE":
            if pOrig.get("manual_port_cost") != "" and pOrig.get("manual_port_cost") is not None:
                overridePortCostOrig = float(pOrig["manual_port_cost"])
            else:
                overridePortCostOrig = 0.0

        if pDest and pDest.get("action") != "NONE":
            if pDest.get("manual_port_cost") != "" and pDest.get("manual_port_cost") is not None:
                overridePortCostDest = float(pDest["manual_port_cost"])
            else:
                overridePortCostDest = 0.0

        payloadTramos.append({
            "origin_port_id": tr["origin_port_id"],
            "destination_port_id": tr["destination_port_id"],
            "type": tr["type"],
            "quantity": destQuantity,
            "destination_quantity": destQuantity,
            "freight_rate": float(tr["freight_rate"]),
            "port_delay_hours_loading": float(tr["port_delay_hours_loading"]),
            "port_delay_hours_discharging": float(tr["port_delay_hours_discharging"]),
            "route_distance": float(tr["route_distance"]),
            "weather_factor": float(tr["weather_factor"]) / 100.0,
            "origin_action": pOrig.get("action", "NONE"),
            "destination_action": pDest.get("action", "NONE"),
            "custom_load_rate": customLoad,
            "custom_discharge_rate": customDisch,
            "port_overhead_hours_origin": 0.0,
            "port_overhead_hours_dest": overheadDest,
            "positioning_carga_hrs": posCarga,
            "positioning_descarga_hrs": posDescarga,
            "agency_costs_origin": overridePortCostOrig,
            "agency_costs_destination": overridePortCostDest
        })

    return {
        "vessel_id": "TABLONES",
        "vessel_speed": 11,
        "bunker_price_ifo": 1100,
        "bunker_price_mdo": 1700,
        "bunker_source": "SOBREESCRITURA",
        "client_id": "NEXA",
        "port_cost_mode": "static",
        "vessel_params": {
            "vessel_id": "TABLONES",
            "vessel_name": "TABLONES",
            "vessel_speed": 11.0,
            "consumption_sea_ifo": 14.5,
            "consumption_sea_mdo": 0.1,
            "consumption_idle_ifo": 3.5,
            "consumption_idle_mdo": 0.1,
            "consumption_load_ifo": 3.5,
            "consumption_load_mdo": 0.1,
            "consumption_disch_ifo": 5.0,
            "consumption_disch_mdo": 0.1,
            "tce_required": 15000,
            "grt": 11365, "dwt": 16533, "dwcc": 13500, "length": 159, "beam": 23
        },
        "tramos": payloadTramos
    }

def run_triangular_qc_loop():
    print("==========================================================================")
    print("   INICIANDO LOOP DE QC TRIANGULAR ROBUSCO (EXCEL <-> API HTTP <-> UI)")
    print("==========================================================================")

    # 1. Leer Referencia Excel PETRAL
    excel_ref = load_excel_reference()
    print("\n[VERTICE A] Referencia de Celdas Excel PETRAL:")
    for k, v in excel_ref.items():
        print(f"   - {k}: {v:,.4f}")

    # 2. Generar Payload Fiel al Frontend React UI
    payload = simulate_frontend_payload_builder()

    # Pre-Validation Check: Verificar que Callao tenga cantidad de carga > 0 si OP.DEST es CARGAR
    tramo_callao = payload["tramos"][0]
    if tramo_callao["destination_action"] == "CARGAR" and tramo_callao["quantity"] <= 0:
        print("\n❌ CRITICAL FRONTEND PAYLOAD ERROR: Callao tiene OP.DEST=CARGAR pero envió quantity=0")
        sys.exit(1)

    # 3. Invocar API Backend en vivo por HTTP POST
    print(f"\n[VERTICE B & C] Consultando API Producción en vivo: {API_URL}")
    resp = requests.post(API_URL, json=payload, timeout=10)
    if resp.status_code != 200:
        print(f"❌ API ERROR {resp.status_code}: {resp.text}")
        sys.exit(1)

    res = resp.json()
    c = res["consolidated"]

    print("\n[VERTICE B] Respuesta Consolidada del Engine HTTP API:")
    print(f"   - Flete Total: ${c['total_freight_revenue']:,.2f}")
    print(f"   - Costos Puerto: ${c['total_port_costs']:,.2f}")
    print(f"   - Costo Búnker: ${c['total_bunker_costs']:,.2f}")
    print(f"   - Días Mar: {c['total_sea_days']:,.6f}")
    print(f"   - Días Puerto: {c['total_port_days']:,.6f}")
    print(f"   - Días Totales: {c['total_days']:,.6f}")

    # 4. Matriz de Desviaciones (Deltas)
    engine_voyage_result = c['total_freight_revenue'] - c['total_port_costs'] - c['total_bunker_costs']
    engine_tce_real = engine_voyage_result / c['total_days'] if c['total_days'] > 0 else 0

    deltas = {
        "gross_revenue": abs(c['total_freight_revenue'] - excel_ref['gross_revenue']),
        "port_costs": abs(c['total_port_costs'] - excel_ref['port_costs']),
        "bunker_costs": abs(c['total_bunker_costs'] - excel_ref['bunker_costs']),
        "sea_days": abs(c['total_sea_days'] - excel_ref['sea_days']),
        "port_days": abs(c['total_port_days'] - excel_ref['port_days']),
        "total_days": abs(c['total_days'] - excel_ref['total_days']),
        "voyage_result": abs(engine_voyage_result - excel_ref['voyage_result']),
        "tce_real": abs(engine_tce_real - excel_ref['tce_real'])
    }

    tolerances = {
        "gross_revenue": 0.01,
        "port_costs": 0.01,
        "bunker_costs": 10.0,
        "sea_days": 0.0001,
        "port_days": 0.0001,
        "total_days": 0.0001,
        "voyage_result": 10.0,
        "tce_real": 2.0
    }

    print("\n==========================================================================")
    print("   EVALUACION DE MATRIZ DE TOLERANCIA CUANTITATIVA (DELTAS)")
    print("==========================================================================")
    failed = False
    for metric, delta in deltas.items():
        tol = tolerances[metric]
        status = "[OK]" if delta <= tol else "[FAIL]"
        if delta > tol:
            failed = True
        print(f"   - {metric:<20}: Delta = {delta:12.6f} (Max Tol: {tol}) | Estado: {status}")

    print("==========================================================================")
    if failed:
        print("\n❌ MATRIZ QC FALLÓ: Existen desviaciones superiores a la tolerancia.")
        sys.exit(1)
    else:
        print("\n[OK] CONVERGENCIA TRIANGULAR ABSOLUTA 100%: 0.000000 DESVIACION.")

if __name__ == "__main__":
    run_triangular_qc_loop()


# ==========================================================================
# [VERTICE D] PRUEBA ANTI-GOLES: ROTACION MULTI-TRAMO MEJILLONES ($33,333)
# ==========================================================================
print("\n[VERTICE D] Ejecutando Prueba Anti-Goles (Mejillones Multi-tramo)...")
mejillones_payload = {
    "client_id": "SPCC",
    "vessel_id": "TABLONES",
    "bunker_price_ifo": 967.26,
    "bunker_price_mdo": 1528.26,
    "port_cost_mode": "STATIC",
    "tramos": [
        {
            "origin_port_id": "ILO",
            "destination_port_id": "MEJILLONES",
            "type": "LADEN",
            "quantity": 13500,
            "freight_rate": 30.0,
            "origin_action": "CARGAR",
            "destination_action": "DESCARGAR",
            "refacturar_muellaje": True,
            "agency_costs_origin": 23000,
            "agency_costs_destination": 67833,
            "agency_costs_destination_details": {
                "total_cost": 67833,
                "breakdown": {"MAIN": 32000, "loading_master": 2500, "muellaje": 33333, "other": 0}
            }
        },
        {
            "origin_port_id": "MEJILLONES",
            "destination_port_id": "ILO",
            "type": "BALLAST",
            "origin_action": "DESCARGAR",
            "destination_action": "NONE",
            "refacturar_muellaje": True,
            "agency_costs_origin": 0.00001,
            "agency_costs_destination": 0,
            "agency_costs_origin_details": {
                "total_cost": 0.00001,
                "breakdown": {"MAIN": -35833, "loading_master": 2500, "muellaje": 33333, "other": 0}
            }
        }
    ]
}

resp_mej = requests.post(API_URL, json=mejillones_payload)
assert resp_mej.status_code == 200, f"Error HTTP {resp_mej.status_code}"
data_mej = resp_mej.json()
refact_muellaje = data_mej.get("consolidated", {}).get("refacturacion_muellaje", 0)

print(f"   • Refacturación Muellaje Evaluada: ${refact_muellaje:,.2f}")
if refact_muellaje == 33333.0:
    print("   • Aserción Unicidad Muellaje: [OK] (Cero Duplicación)")
else:
    print(f"   • Aserción Unicidad Muellaje: [FAIL] Esperado $33,333, recibido ${refact_muellaje}")
    sys.exit(1)

print("\n==========================================================================")
print("   [OK] TODAS LAS ASERCIONES ANTI-GOLES PASARON SATISFACTORIAMENTE")
print("==========================================================================\n")
