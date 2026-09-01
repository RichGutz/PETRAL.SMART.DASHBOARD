import openpyxl
import json

excel_path = 'C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/test_qc_matriz_financiera_verified.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb.active

print(f'=== REPORTE DE AUDITORÍA PERICIAL QC (BENOIT BLANC) ===')
print(f'Archivo: {excel_path}')
print(f'Pestaña: {ws.title}')
print(f'Dimensiones: {ws.max_row} Filas x {ws.max_column} Columnas')
print(f'Total Rangos Combinados (Merged Cells): {len(ws.merged_cells.ranges)}')
print(f'Rangos Combinados:', [str(r) for r in list(ws.merged_cells.ranges)[:8]])

print('\n[1. CABECERAS DE COLUMNA - THEAD]')
headers = []
for col in range(1, 18):
    cell = ws.cell(row=1, column=col)
    fill_hex = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
    font_color = cell.font.color.rgb if cell.font and cell.font.color else 'None'
    headers.append(f'{cell.coordinate}: {cell.value} (Fill: {fill_hex}, Font: {font_color})')
for h in headers:
    print(' ', h)

print('\n[2. COLUMNAS DE DIMENSIÓN Y ROTACIÓN VERTICAL A 90°]')
for col, name in [(1, 'CLIENTE'), (2, 'RUTA'), (3, 'BUQUE')]:
    cell = ws.cell(row=2, column=col)
    fill_hex = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
    font_color = cell.font.color.rgb if cell.font and cell.font.color else 'None'
    rotation = cell.alignment.textRotation if cell.alignment else 'None'
    print(f' {name:<8} ({cell.coordinate}): Texto="{cell.value}" | Rotación={rotation}° | Fondo={fill_hex} | TextoColor={font_color} | Bold={cell.font.bold}')

print('\n[3. CONTRASTE DE CELDAS DE DATOS Y FORMATOS NUMÉRICOS EXCEL]')
for row in range(2, min(ws.max_row + 1, 15)):
    m_cell = ws.cell(row=row, column=4)
    ene_cell = ws.cell(row=row, column=5)
    tot_cell = ws.cell(row=row, column=17)
    print(f' Fila {row:2d} | Métrica: {str(m_cell.value):<24} | Ene: {ene_cell.value!s:<10} [Fmt: {ene_cell.number_format}] | Tot Acum: {tot_cell.value!s:<10} [Fmt: {tot_cell.number_format}]')

print('\n[4. VALIDACIÓN DE NO-CONCATENACIÓN DE SELECTORES]')
vessel_cells = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=3).value]
print(' Buques identificados en celdas:', set(vessel_cells))
has_concat = any(len(v) > 25 for v in vessel_cells if v)
print(' ¿Existe concatenación parásita de selects?:', 'SÍ (ERROR)' if has_concat else 'NO (CORRECTO - 100% LIMPIO)')
