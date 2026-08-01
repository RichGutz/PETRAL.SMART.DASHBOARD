import os
import sys
import json
import datetime
import openpyxl

file_jn = r'C:\Users\rguti\PETRAL.SMART.DASHBOARD\Documentos.Petral\Resultados.JN\VC Tablones 2026.PASS.THROUGH.xlsx'
file_mec = r'C:\Users\rguti\PETRAL.SMART.DASHBOARD\Documentos.Petral\Resultados.MEC\MOQUEGUA - Voyage calculation viajes Enero a Junio  2026  - 31.07.2026.PASS.THROUGH.xlsx'

def clean_val(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d %H:%M')
    return val

def parse_jn_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    # Check if valid voyage sheet
    vessel = ws.cell(row=5, column=3).value
    if not vessel or 'TABLONES' not in str(vessel).upper():
        return None

    prepared_by = ws.cell(row=1, column=8).value or 'JN'
    dwt = ws.cell(row=5, column=7).value or 16500

    # Income & Granular Freight Breakdown (B15:B22, I15:I22, Total I23)
    freight_income_items = []
    for r in range(15, 23):
        concept = ws.cell(row=r, column=2).value
        val = ws.cell(row=r, column=9).value
        if concept and str(concept).strip() and val is not None:
            try:
                amount = float(val)
                freight_income_items.append({"concept": str(concept).strip(), "amount_usd": amount})
            except (ValueError, TypeError):
                pass

    cargo_charterer = ws.cell(row=15, column=2).value or 'SPCC'
    freight_rate = float(ws.cell(row=15, column=3).value or 0)
    qty_mt = float(ws.cell(row=15, column=5).value or 0)
    gross_rev = float(ws.cell(row=23, column=9).value or ws.cell(row=15, column=7).value or ws.cell(row=14, column=14).value or 0)
    if gross_rev == 0 and qty_mt > 0 and freight_rate > 0:
        gross_rev = qty_mt * freight_rate

    # Voyage Program / Itinerary
    itinerary = []
    pol_port = 'ILO'
    pod_port = 'MEJILLONES'
    stops = []

    for r in range(29, 36):
        port_name = ws.cell(row=r, column=2).value
        if port_name and str(port_name).strip() and str(port_name).strip().upper() != 'TOTAL':
            p_clean = str(port_name).strip().upper()
            arr = clean_val(ws.cell(row=r, column=4).value)
            qty = ws.cell(row=r, column=7).value or 0
            rate = ws.cell(row=r, column=8).value or 0
            other_h = ws.cell(row=r, column=10).value or 0
            itinerary.append({
                "port": p_clean,
                "arrival_datetime": arr,
                "quantity_mt": qty,
                "rate_mth": rate,
                "other_hrs": other_h
            })
            stops.append(p_clean)

    pols = []
    pods = []
    for item in itinerary:
        try:
            q = float(item.get("quantity_mt") or 0.0)
        except (ValueError, TypeError):
            q = 0.0
        if q > 0:
            pols.append(item["port"])
        elif q < 0:
            pods.append(item["port"])

    if pols:
        pol_port = " / ".join(pols)
    elif len(stops) > 0:
        pol_port = stops[0]
    else:
        pol_port = "ILO"

    if pods:
        pod_port = " / ".join(pods)
    elif len(stops) > 0:
        pod_port = stops[-1]
    else:
        pod_port = "MEJILLONES"


    # Port & Bunker Expenses (Lectura exacta N15/N16 + C48/S48)
    agency_cost = float(ws.cell(row=15, column=14).value or ws.cell(row=48, column=3).value or ws.cell(row=48, column=4).value or 0)
    bunker_cost = float(ws.cell(row=16, column=14).value or ws.cell(row=48, column=19).value or 0)

    # Consumption & Duration (Estimates)
    sea_days = float(ws.cell(row=5, column=19).value or ws.cell(row=15, column=17).value or 0)
    tce_usd = float(ws.cell(row=17, column=17).value or 0)
    tce_req = float(ws.cell(row=18, column=17).value or 13000.0)
    net_profit = float(ws.cell(row=20, column=17).value or 0)

    client_name = 'NEXA' if 'NEXA' in str(cargo_charterer).upper() or 'NEXA' in sheet_name.upper() else 'SPCC'

    details = {
        "vessel_header": {
            "vessel_name": "B/T TABLONES",
            "prepared_by": str(prepared_by),
            "dwt": dwt
        },
        "income": {
            "charterer": client_name,
            "cargo_name": str(cargo_charterer),
            "freight_rate_usd": freight_rate,
            "quantity_mt": qty_mt,
            "gross_revenue_usd": gross_rev,
            "freight_income_items": freight_income_items,
            "total_freight_income_usd": gross_rev
        },
        "itinerary": itinerary,
        "port_expenses": {
            "total_agency_usd": agency_cost
        },
        "bunker_expenses": {
            "total_bunker_cost_usd": bunker_cost
        }
    }

    clean_vcode = sheet_name.replace(' (2)', '').replace('(2)', '').strip()
    return {
        "voyage_code": clean_vcode,
        "vessel_name": "B/T Tablones",

        "client_name": client_name,
        "voyage_date": "2026-01-15",
        "pol_port": pol_port,
        "pod_port": pod_port,
        "stops": json.dumps(stops),
        "operator": "JN",
        "cargo_quantity_mt": qty_mt,
        "freight_rate_usd": freight_rate,
        "gross_revenue_usd": gross_rev,
        "tce_usd_day": tce_usd,
        "tce_req_usd_day": tce_req,
        "pcm_usd": gross_rev - bunker_cost - agency_cost,
        "net_profit_usd": net_profit if net_profit != 0 else (gross_rev - bunker_cost - agency_cost),
        "details": json.dumps(details)
    }

def parse_mec_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    vessel = ws.cell(row=5, column=3).value
    if not vessel or 'MOQUEGUA' not in str(vessel).upper():
        return None

    prepared_by = ws.cell(row=1, column=8).value or 'MEC'
    dwt = ws.cell(row=5, column=7).value or 14300

    # Income & Granular Freight Breakdown (B15:B22, I15:I22, Total I23)
    freight_income_items = []
    for r in range(15, 23):
        concept = ws.cell(row=r, column=2).value
        val = ws.cell(row=r, column=9).value
        if concept and str(concept).strip() and val is not None:
            try:
                amount = float(val)
                freight_income_items.append({"concept": str(concept).strip(), "amount_usd": amount})
            except (ValueError, TypeError):
                pass

    cargo_charterer = ws.cell(row=15, column=2).value or 'SPCC'
    freight_rate = float(ws.cell(row=15, column=3).value or 0)
    qty_mt = float(ws.cell(row=15, column=5).value or 0)
    gross_rev = float(ws.cell(row=23, column=9).value or ws.cell(row=23, column=8).value or ws.cell(row=14, column=14).value or ws.cell(row=15, column=7).value or 0)
    if gross_rev == 0 and qty_mt > 0 and freight_rate > 0:
        gross_rev = qty_mt * freight_rate

    # KPIs block on right (Cols 14-20)
    tce_usd = float(ws.cell(row=17, column=17).value or 0)
    tce_req = float(ws.cell(row=18, column=17).value or 13000)
    pcm_val = float(ws.cell(row=19, column=17).value or 0)
    pnl_val = float(ws.cell(row=20, column=17).value or 0)

    # Voyage Program / Itinerary
    itinerary = []
    pol_port = 'ILO'
    pod_port = 'MARCONA'
    stops = []

    for r in range(29, 36):
        port_name = ws.cell(row=r, column=2).value
        if port_name and str(port_name).strip() and str(port_name).strip().upper() != 'TOTAL':
            p_clean = str(port_name).strip().upper()
            arr = clean_val(ws.cell(row=r, column=4).value)
            qty = ws.cell(row=r, column=7).value or 0
            rate = ws.cell(row=r, column=8).value or 0
            other_h = ws.cell(row=r, column=10).value or 0
            dist = ws.cell(row=r, column=16).value or 0
            speed = ws.cell(row=r, column=17).value or 11
            sailing_h = ws.cell(row=r, column=18).value or 0
            itinerary.append({
                "port": p_clean,
                "arrival_datetime": arr,
                "quantity_mt": qty,
                "rate_mth": rate,
                "other_hrs": other_h,
                "dist_nm": dist,
                "speed_kts": speed,
                "sailing_hrs": sailing_h
            })
            stops.append(p_clean)

    pols = []
    pods = []
    for item in itinerary:
        try:
            q = float(item.get("quantity_mt") or 0.0)
        except (ValueError, TypeError):
            q = 0.0
        if q > 0:
            pols.append(item["port"])
        elif q < 0:
            pods.append(item["port"])

    if pols:
        pol_port = " / ".join(pols)
    elif len(stops) > 0:
        pol_port = stops[0]
    else:
        pol_port = "ILO"

    if pods:
        pod_port = " / ".join(pods)
    elif len(stops) > 0:
        pod_port = stops[-1]
    else:
        pod_port = "MARCONA"


    # Port & Bunker Expenses (Lectura exacta N15/N16 + C48/S48)
    agency_cost = float(ws.cell(row=15, column=14).value or ws.cell(row=48, column=3).value or ws.cell(row=48, column=4).value or 0)
    bunker_cost = float(ws.cell(row=16, column=14).value or ws.cell(row=48, column=19).value or 0)

    client_name = 'NEXA' if 'NEXA' in str(cargo_charterer).upper() or 'NEXA' in sheet_name.upper() else 'SPCC'

    details = {
        "vessel_header": {
            "vessel_name": "B/T MOQUEGUA",
            "prepared_by": str(prepared_by),
            "dwt": dwt
        },
        "income": {
            "charterer": client_name,
            "cargo_name": str(cargo_charterer),
            "freight_rate_usd": freight_rate,
            "quantity_mt": qty_mt,
            "gross_revenue_usd": gross_rev,
            "freight_income_items": freight_income_items,
            "total_freight_income_usd": gross_rev
        },
        "itinerary": itinerary,
        "port_expenses": {
            "total_agency_usd": agency_cost
        },
        "bunker_expenses": {
            "total_bunker_cost_usd": bunker_cost
        }
    }

    clean_vcode = sheet_name.replace(' (2)', '').replace('(2)', '').strip()
    return {
        "voyage_code": clean_vcode,
        "vessel_name": "B/T Moquegua",

        "client_name": client_name,
        "voyage_date": "2026-01-15",
        "pol_port": pol_port,
        "pod_port": pod_port,
        "stops": json.dumps(stops),
        "operator": "MEC",
        "cargo_quantity_mt": qty_mt,
        "freight_rate_usd": freight_rate,
        "gross_revenue_usd": gross_rev,
        "tce_usd_day": tce_usd,
        "tce_req_usd_day": tce_req,
        "pcm_usd": pcm_val if pcm_val != 0 else (gross_rev - bunker_cost - agency_cost),
        "net_profit_usd": pnl_val if pnl_val != 0 else (gross_rev - bunker_cost - agency_cost),
        "details": json.dumps(details)
    }

