import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

sys.stdout.reconfigure(encoding='utf-8')

def generate_matrix():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz Fórmulas Multicotizador"

    # Paleta de colores
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Azul Petral
    total_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Azul Total
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="1E293B")
    font_regular = Font(name="Calibri", size=10, color="334155")
    font_code = Font(name="Consolas", size=9.5, color="0F172A")
    font_total = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # Título
    ws.cell(row=2, column=2, value="📊 PETRAL SMART DASHBOARD — ESPECIFICACIÓN Y MATRIZ MATEMÁTICA DE LA GRILLA").font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    ws.cell(row=3, column=2, value="Mapeo columna por columna, tablas de origen en Supabase y fórmulas en pseudocódigo").font = Font(name="Calibri", size=10, italic=True, color="64748B")

    headers = [
        "COL",
        "NOMBRE COLUMNA",
        "TIPO CAMPO",
        "ORIGEN INSUMO",
        "TABLA & KEY SUPABASE",
        "FÓRMULA / PSEUDOCÓDIGO (EN PALABRAS SIMPLES)",
        "EJEMPLO CASO REAL (ILO ➔ MATARANI ➔ ILO)"
    ]

    for col_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    rows_data = [
        ("A", "LEG", "Correlativo", "Estructura UI", "—", "SI Fila == 0 ENTONCES '—' SINO index_fila", "Fila 0 = '—', Fila 1 = 1, Fila 2 = 2"),
        ("B", "TIPO", "Badge Estado", "Motor Matemático", "—", "SI Carga_a_Bordo > 0 ENTONCES 'LADEN' SINO 'BALLAST'", "Tramo 1 = LADEN, Tramo 2 = BALLAST"),
        ("C", "PUERTO", "Dropdown", "Catálogo BD", "Tabla: ports (port_id, port_name)", "Selección del usuario desde catálogo oficial de puertos", "Fila 0 = ILO, Fila 1 = MATARANI, Fila 2 = ILO"),
        ("D", "DIST (NM)", "Numérico Editable", "Matriz BD + Input", "Tabla: distances (port_a, port_b, route_distance)", "Buscar en distances (port_a=Origen AND port_b=Destino). Si usuario edita, sobreescribir.", "69 NM (ILO-MATARANI), 69 NM (MATARANI-ILO)"),
        ("E", "W.F (%)", "Numérico Editable", "Matriz BD + Input", "Tabla: distances (weather_factor_laden / ballast)", "SI valor_bd <= 1.0 ENTONCES valor_bd * 100 SINO valor_bd. Fallback = 3.0%.", "3.0%"),
        ("F", "VEL (KN)", "Numérico Editable", "Buque BD + Input", "Tabla: vessels (vessel_id, vessel_speed)", "Velocidad estándar del buque seleccionado. Si usuario edita, propaga a todos los tramos.", "11.0 kn"),
        ("G", "DÍAS MAR", "Cálculo Solo Lectura", "Motor Matemático", "—", "FÓRMULA: (DIST * (1 + WF/100)) / (VEL * 24). En Fila 0 es '—'.", "Tramo 1: (69 * 1.03) / (11 * 24) = 0.27 d\nTramo 2: (69 * 1.03) / (11 * 24) = 0.27 d"),
        ("H", "DÍAS PTO", "Cálculo Solo Lectura", "Motor Matemático", "—", "FÓRMULA: Días_Espera + Días_Operación = ((TTC + Posic) / 24) + ((Q / Ritmo) / FactorUnidad)", "Fila 0 = 1.83 d (17h idle + 27h op)\nFila 1 = 1.54 d (7h idle + 30h op)\nFila 2 = 0.00 d"),
        ("I", "TIME TO COUNT (H)", "Input + Placeholder", "Input Usuario + Regla", "Regla Petral (Fallback = 6.0)", "SI usuario digita valor ENTONCES valor SINO sugerir gris '6.0' (tanto en Carga como Descarga)", "Fila 0 = 7.0 h, Fila 1 = 7.0 h"),
        ("J", "POSIC (H)", "Input + Placeholder", "Input Usuario + Regla", "Regla Petral (1.0 Carga / 0.0 Descarga)", "SI usuario digita valor ENTONCES valor SINO (SI Accion=='CARGAR' ENTONCES '1.0' SINO '0.0')", "Fila 0 = 10.0 h (Carga), Fila 1 = 0.0 h (Descarga)"),
        ("K", "OP. DEST", "Dropdown Selector", "Decisión Operador", "—", "Selector de acción operativa: 'CARGAR', 'DESCARGAR', 'NONE'", "Fila 0 = CARGAR, Fila 1 = DESCARGAR, Fila 2 = NONE"),
        ("L", "RITMO (C/D)", "Input + Unidad", "Input Usuario + Regla", "Regla Petral (500 Carga / 450 Descarga)", "SI usuario digita ritmo ENTONCES ritmo SINO sugerir gris (500 en Carga, 450 en Descarga). Unidad T/h o T/d.", "Fila 0 = 500 T/h, Fila 1 = 450 T/h"),
        ("M", "Q (MT)", "Numérico Editable", "Input Usuario", "—", "Cantidad de toneladas ingresada por el usuario para cargar o descargar", "Fila 0 = 13,500 MT, Fila 1 = 13,500 MT"),
        ("N", "F ($/T)", "Numérico Editable", "Input Usuario", "—", "Tarifa de flete en $/MT ingresada en tramos de DESCARGAR", "Fila 1 = $20.00 / MT"),
        ("O", "COSTO PTO ($)", "Numérico Editable", "Tarifario BD + Input", "Tabla: port_cost_static / port_costs_matrix", "Buscar costo portuario por (puerto, buque, operacion). Editable por el usuario.", "Fila 0 = $23,000, Fila 1 = $22,000"),
        ("P", "FLETE ($)", "Cálculo Solo Lectura", "Motor Matemático", "—", "SI Accion == 'DESCARGAR' ENTONCES Q * F SINO $0", "Fila 1: 13,500 * $20 = $270,000"),
        ("Q", "BUNKER ($)", "Cálculo Solo Lectura", "Motor Matemático", "Tablas: bunker_prices Y vessels", "FÓRMULA: (Total_Tons_IFO * Precio_IFO) + (Total_Tons_MDO * Precio_MDO). Incluye Mar + Espera + Operación.", "Fila 0 = $6,487\nFila 1 = $11,085\nFila 2 = $3,817"),
        ("R", "MUELLAJE ($)", "Cálculo / Input", "Matriz BD + Input", "Tabla: port_costs_matrix (allow_pass_through=true)", "Costo de muellaje parametrizado (ej. Mejillones $33,333 o gasto local)", "Fila 1 = $4,000"),
        ("S", "RF (Checkbox)", "Checkbox Booleano", "Decisión Comercial", "—", "SI [x] Marcado ENTONCES Refactura muellaje al cliente (suma a Gross Revenue) SINO Armador lo absorbe.", "[x] Marcado = True")
    ]

    current_row = 6
    for idx, row in enumerate(rows_data):
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = border_thin
            if idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in [2, 3]:
                cell.font = font_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:
                cell.font = font_regular
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [5, 6]:
                cell.font = font_code
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 7:
                cell.font = font_code
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.font = font_regular
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        current_row += 1

    # Fila TOTAL
    ws.append([])
    current_row += 1
    total_cells = [
        ("B", "📊 TOTAL AZUL"),
        ("C", "Resumen Viaje"),
        ("D", "Motor Central"),
        ("E", "liveCalculation"),
        ("F", "SUMA VERTICAL PURA: Distancia = sum(Dist), Días Mar = sum(Días_Mar), Días Pto = sum(Días_Pto), Flete = sum(Flete), Búnker = sum(Búnker)"),
        ("G", "Dist: 138 NM | Días Mar: 0.54 d | Días Pto: 3.38 d | Flete: $270,000 | Búnker: $21,389")
    ]
    
    cell_tag = ws.cell(row=current_row, column=2, value="TOTAL")
    cell_tag.fill = total_fill
    cell_tag.font = font_total
    cell_tag.alignment = Alignment(horizontal="center", vertical="center")
    cell_tag.border = border_thin

    for col_idx, (col_letter, text) in enumerate(total_cells, start=3):
        cell = ws.cell(row=current_row, column=col_idx, value=text)
        cell.fill = total_fill
        cell.font = font_total
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left" if col_idx in [6, 7, 8] else "center", vertical="center", wrap_text=True)

    # Anchos de columna
    col_widths = {
        1: 4,
        2: 8,
        3: 22,
        4: 20,
        5: 25,
        6: 35,
        7: 60,
        8: 40
    }
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    output_path = r"C:\Users\rguti\PETRAL.SMART.DASHBOARD\Exceles.Petral\MATRIZ_FORMULAS_MULTICOTIZADOR_EXCEL.xlsx"
    wb.save(output_path)
    print(f"✅ Archivo Excel generado exitosamente en: {output_path}")

if __name__ == "__main__":
    generate_matrix()
