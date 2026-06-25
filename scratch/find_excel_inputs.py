import openpyxl

files = [
    "C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Tablones.xlsx",
    "C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Moquegua.xlsx",
    "C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Concon_Trader.xlsx"
]

def search_inputs(ws):
    found = {}
    for row in ws.iter_rows(values_only=True):
        if not row: continue
        for i, cell in enumerate(row):
            if cell and isinstance(cell, str):
                val = str(cell).lower().strip()
                # Buscar palabras clave comunes para inputs
                if any(x in val for x in ["sea days", "port days", "freight income", "total agency"]):
                    for j in range(1, 4):
                        if i + j < len(row):
                            adj = row[i+j]
                            if isinstance(adj, (int, float)):
                                found[val] = adj
                                break
    return found

for f in files:
    print(f"\nArchivo: {f.split('/')[-1]}")
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            inputs = search_inputs(ws)
            print(f"  Hoja: {sheet_name}")
            for k, v in inputs.items():
                print(f"    {k}: {v}")
    except Exception as e:
        print(f"Error reading {f}: {e}")
