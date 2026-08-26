import openpyxl

excel_path = r"C:\Users\rguti\PETRAL.SMART.DASHBOARD\Documentos.Petral\Control_Presupuestal_NAVITRANSO.xlsx"

wb = openpyxl.load_workbook(excel_path, data_only=False) # Con formulas
ws = wb['Control Presupuestal']

print("=== ESTRUCTURA Y FÓRMULAS OFICIALES NAVITRANS ===")
for r in range(1, ws.max_row + 1):
    row_data = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None:
            row_data.append(f"C{c}: {v}")
    if row_data:
        print(f"Fila {r:2d} -> {' | '.join(row_data)}")
