import os
import re
import psycopg2
import openpyxl

URI = "postgresql://postgres.hjjxooxcpvlvbaxgifbn:VivaLaVida2026$@aws-1-us-east-2.pooler.supabase.com:6543/postgres"
FOLDER = r"C:\Users\rguti\PETRAL.SMART.DASHBOARD\Exceles.Petral\PORT.COSTS.PATRICIA"

FILE_MAP = {
    "CALLAO": "Costos.SUA  %  2026 01.07.2026 REV AL 03.07.CALLAO.xlsx",
    "ILO": "Costos.SUA  %  2026 01.07.2026 REV AL 07.07.ILO.xlsx",
    "MATARANI": "Costos.SUA  %  2026 01.07.2026 REV AL 07.07.MATARANI.xlsx",
    "MARCONA": "Costos.SUA  %  2026 01.07.2026 REV AL 08.07.MARCONA.xlsx",
}
CHILE_FILE = "Costos Moquegua SUA  %  2026 01.07.2026 REV AL 13.07 TPM MEJILLONES, INTERACID,TERQUIM,BARQUITO.xlsx"

def extract_times(formula_str):
    if not formula_str or not isinstance(formula_str, str):
        return 0.0, 0.0
    # Match something like +3+2 or + 3 + 2 or +3.5+2
    match = re.search(r'\+\s*(\d+(?:\.\d+)?)\s*\+\s*(\d+(?:\.\d+)?)', formula_str)
    if match:
        return float(match.group(1)), float(match.group(2))
    return 0.0, 0.0

def main():
    conn = psycopg2.connect(URI)
    cur = conn.cursor()
    
    # Obtener terminales actuales
    cur.execute("SELECT terminal_id, port_id FROM terminals")
    terminals = cur.fetchall()
    
    updates = []
    
    # 1. Procesar puertos de Perú
    for port_id, filename in FILE_MAP.items():
        filepath = os.path.join(FOLDER, filename)
        if not os.path.exists(filepath):
            print(f"Skipping {port_id}: File not found.")
            continue
            
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb.active
        
        formula = ws.cell(row=11, column=6).value
        if not isinstance(formula, str):
            formula = ws.cell(row=11, column=8).value # Check fallback if col F is empty
            
        mooring, unmooring = extract_times(formula)
        print(f"[{port_id}] Peru -> Found formula: {formula} -> Mooring: {mooring}, Unmooring: {unmooring}")
        
        for t_id, p_id in terminals:
            if p_id == port_id:
                updates.append((mooring, unmooring, t_id, p_id))

    # 2. Procesar puertos de Chile (Múltiples hojas)
    chile_path = os.path.join(FOLDER, CHILE_FILE)
    if os.path.exists(chile_path):
        wb_chile = openpyxl.load_workbook(chile_path, data_only=False)
        
        # Mapeo de Terminales de Chile
        # port_id -> (sheet_name, is_specific_terminal_mapping)
        chile_mapping = {
            "MEJILLONES": {
                "TERMINAL_A": "MEJILLONES", 
                "TPM": "MEJILLONES",
                "INTERACID": "MEJILLONES INTERACID",
                "TERQUIM": "TERQUIM"
            },
            "BARQUITO": {
                "GENERAL": "BARQUITO"
            }
        }
        
        for t_id, p_id in terminals:
            if p_id in chile_mapping:
                sheet_name = None
                if isinstance(chile_mapping[p_id], dict):
                    sheet_name = chile_mapping[p_id].get(t_id)
                
                if sheet_name and sheet_name in wb_chile.sheetnames:
                    ws = wb_chile[sheet_name]
                    formula = ws.cell(row=11, column=6).value
                    if not isinstance(formula, str):
                        formula = ws.cell(row=11, column=8).value
                        
                    mooring, unmooring = extract_times(formula)
                    print(f"[{p_id} - {t_id}] Chile -> Sheet: {sheet_name} | Formula: {formula} -> Mooring: {mooring}, Unmooring: {unmooring}")
                    updates.append((mooring, unmooring, t_id, p_id))
    else:
        print("Chile file not found!")
        
    # Aplicar Updates
    print("\n--- Ejecutando UPDATES en DB ---")
    for moor, unmoor, t_id, p_id in updates:
        cur.execute("""
            UPDATE terminals 
            SET mooring_time_hrs = %s, unmooring_time_hrs = %s 
            WHERE terminal_id = %s AND port_id = %s
        """, (moor, unmoor, t_id, p_id))
        print(f"Updated Terminal {t_id} at {p_id}: Amarre={moor}h, Desamarre={unmoor}h")
        
    conn.commit()
    cur.close()
    conn.close()
    print("\nSeeding completado con éxito.")

if __name__ == "__main__":
    main()
