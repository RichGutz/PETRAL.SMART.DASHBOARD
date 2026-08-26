import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\rguti\PETRAL.SMART.DASHBOARD\Documentos.Petral\Control_Presupuestal_NAVITRANSO.xlsx', data_only=False)
ws = wb['Control Presupuestal']

print("=== VERIFICACIÓN DE FÓRMULAS EXACTAS EN EXCEL ===")
for col in ['C', 'D', 'E', 'F', 'G']:
    mes = ws[f'{col}4'].value
    ventas = ws[f'{col}5'].value
    hire = ws[f'{col}6'].value
    demoras = ws[f'{col}8'].value
    muellaje = ws[f'{col}9'].value
    
    costos = ws[f'{col}11'].value
    bunker = ws[f'{col}12'].value
    puerto = ws[f'{col}13'].value
    costo_demora = ws[f'{col}14'].value
    
    tce = ws[f'{col}17'].value
    arriendo = ws[f'{col}18'].value
    margen = ws[f'{col}19'].value
    
    print(f"\n--- MES {mes} ---")
    print(f"  Fila 5  VENTAS = {ventas} (Hire: {hire} + Demoras: {demoras} + Muellaje: {muellaje})")
    print(f"  Fila 11 COSTOS DIRECTOS = {costos} (Bunker: {bunker} + Gastos Pto: {puerto} + Costo Demora: {costo_demora})")
    print(f"  Fila 17 TIME CHARTER EQUIVALENT = {tce}")
    print(f"  Fila 18 COSTO ARRIENDO NAVES = {arriendo}")
    print(f"  Fila 19 MARGEN BRUTO = {margen}")
