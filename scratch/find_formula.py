import openpyxl

f = "C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Tablones.xlsx"

print(f"Abriendo {f}")
wb = openpyxl.load_workbook(f, data_only=False) # Para leer fórmulas
ws = wb["ILO-MATARANI"]

for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            val = str(cell.value).lower().strip()
            if "total bunker cons. (incl. cleaning & heating)" in val:
                print(f"Encontrado texto en celda {cell.coordinate}")
                # Imprimir las celdas de las que depende: O49, M49, C60, C61
                print("Dependencias de la formula =+O49+M49+C60+C61:")
                for c in ["O49", "M49", "C60", "C61"]:
                    print(f"  {c}: {ws[c].value}")
                
                # Para entender el contexto, vamos a ver los textos en la columna M para esas filas
                print("Contexto de las filas:")
                print(f"  M49: {ws['M49'].value}")
                print(f"  B60: {ws['B60'].value} (suponiendo que B tiene el texto de C60)")
                print(f"  B61: {ws['B61'].value} (suponiendo que B tiene el texto de C61)")
                
                break
