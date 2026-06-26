import openpyxl

files = [
    'C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Tablones.xlsx',
    'C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Moquegua.xlsx',
    'C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/Voyage_Calculations_Concon_Trader.xlsx'
]
for f in files:
    print(f"\nReading {f.split('/')[-1]}...")
    wb = openpyxl.load_workbook(f, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name == 'RESUMEN': continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if cell and 'additional expenses' in str(cell).strip().lower():
                    print(f'Sheet {sheet_name}: {row[i+1]}')
