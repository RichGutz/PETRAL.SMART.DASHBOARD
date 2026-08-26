import openpyxl
import pandas as pd
import json

excel_path = r"C:\Users\rguti\PETRAL.SMART.DASHBOARD\Documentos.Petral\Control_Presupuestal_NAVITRANSO.xlsx"

wb = openpyxl.load_workbook(excel_path, data_only=True)
print(f"Hojas en el libro: {wb.sheetnames}\n")

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    max_r = ws.max_row
    max_c = ws.max_column
    print(f"=== HOJA: '{sheetname}' ({max_r} filas x {max_c} columnas) ===")
    
    # Leer primeras 35 filas con contenido no vacío
    rows_data = []
    for r in range(1, min(max_r + 1, 40)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(max_c + 1, 20))]
        if any(v is not None for v in row_vals):
            rows_data.append((r, [str(v) if v is not None else "" for v in row_vals]))
            
    for r_idx, vals in rows_data[:25]:
        non_empty = [f"Col{c+1}: {v}" for c, v in enumerate(vals) if v != ""]
        if non_empty:
            print(f"  Fila {r_idx:2d}: {' | '.join(non_empty[:8])}")
    print("\n" + "-"*60 + "\n")
