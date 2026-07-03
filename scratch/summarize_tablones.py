import openpyxl
import pandas as pd

excel_path = r"C:\Users\rguti\PETRAL.SMART.DASHBOARD\Exceles.Petral\PORT.COSTS.SANDRA\Costos Tablones.01.07.2026.xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True)
print("Hojas:", wb.sheetnames)

for sheet in wb.sheetnames:
    print(f"\n--- Resumen de hoja: {sheet} ---")
    df = pd.read_excel(excel_path, sheet_name=sheet)
    # Filtrar filas que tengan algo de texto interesante
    # Mostrar las primeras 15 filas no vacías
    valid_rows = df.dropna(how='all')
    print(valid_rows.head(25).to_string())
