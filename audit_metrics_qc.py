import openpyxl

excel_path = 'C:/Users/rguti/PETRAL.SMART.DASHBOARD/Exceles.Petral/QC_Auditoria_Escenarios/Matriz_PB_2027_Jose_de_los_Heros__Prom_Dem__NexaRG.xlsx'
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

empty_metrics = []
wrong_monetary = []

print('=== AUDITORÍA PERICIAL DE COLUMNA MÉTRICA Y FORMATOS (96 FILAS) ===')
for r in range(2, ws.max_row + 1):
    m_val = str(ws.cell(row=r, column=4).value or '').strip()
    c5 = ws.cell(row=r, column=5)
    c17 = ws.cell(row=r, column=17)
    
    if not m_val:
        empty_metrics.append(r)
    
    is_non_monetary = any(k in m_val.lower() for k in ['viaje', 'días', 'dias', 'tonelada', 'margen %', 'yield %'])
    if is_non_monetary and ('$' in str(c5.number_format) or '$' in str(c17.number_format)):
        wrong_monetary.append((r, m_val, c5.number_format))

print(f'Total filas evaluadas: {ws.max_row - 1}')
print(f'Filas con columna Métrica vacía: {len(empty_metrics)} -> {empty_metrics}')
print(f'Métricas no monetarias con formato dólar: {len(wrong_monetary)} -> {wrong_monetary}')

print('\nDetalle representativo de todas las filas y sus formatos:')
for r in range(2, min(ws.max_row + 1, 30)):
    m_val = ws.cell(row=r, column=4).value
    c5 = ws.cell(row=r, column=5)
    c17 = ws.cell(row=r, column=17)
    print(f' Fila {r:2d} | Métrica: {str(m_val):<26} | Ene: {c5.value!s:<8} (Fmt: {c5.number_format:<10}) | Tot: {c17.value!s:<10} (Fmt: {c17.number_format})')
