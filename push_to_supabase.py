import os
import requests
import json
import openpyxl
from etl_parser_liquidations import parse_jn_sheet, parse_mec_sheet, file_jn, file_mec


SUPABASE_URL = "https://hjjxooxcpvlvbaxgifbn.supabase.co"
SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imhqanhvb3hjcHZsdmJheGdpZmJuIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MjI1MDk0NCwiZXhwIjoyMDk3ODI2OTQ0fQ.i8KkZtLSDEqaNo15NH3easZV6vhHIbqoYD7ps4pkOMc"

headers = {
    "apikey": SERVICE_ROLE_KEY,
    "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates"
}

def main():
    # 1. Limpiar la tabla en Supabase para asegurar conteo exacto de viajes reales
    del_url = f"{SUPABASE_URL}/rest/v1/voyage_liquidations?id=gt.00000000-0000-0000-0000-000000000000"
    requests.delete(del_url, headers=headers)

    records = []
    
    # 2. Cargar viajes reales de Tablones (15 viajes)
    wb_jn = openpyxl.load_workbook(file_jn, data_only=True)
    for sh in wb_jn.sheetnames:
        if sh.lower().startswith('v.'):
            try:
                rec = parse_jn_sheet(wb_jn, sh)
                if rec:
                    rec['stops'] = json.loads(rec['stops'])
                    rec['details'] = json.loads(rec['details'])
                    records.append(rec)
            except Exception as e:
                print(f"Error procesando {sh}: {e}")

    # 3. Cargar viajes reales de Moquegua (16 viajes)
    wb_mec = openpyxl.load_workbook(file_mec, data_only=True)
    for sh in wb_mec.sheetnames:
        if sh.lower().startswith('v.'):
            try:
                rec = parse_mec_sheet(wb_mec, sh)
                if rec:
                    rec['stops'] = json.loads(rec['stops'])
                    rec['details'] = json.loads(rec['details'])
                    records.append(rec)
            except Exception as e:
                print(f"Error procesando {sh}: {e}")

    print(f"Upsertando exactamente {len(records)} viajes reales en la tabla 'voyage_liquidations' de Supabase...")
    
    url = f"{SUPABASE_URL}/rest/v1/voyage_liquidations"
    res = requests.post(url, headers=headers, json=records)
    
    if res.status_code in [200, 201]:
        print(f"EXITO TOTAL: Se upsertaron los {len(records)} viajes reales correctamente en Supabase!")
    else:
        print(f"Respuesta Supabase [{res.status_code}]: {res.text}")

if __name__ == '__main__':
    main()

